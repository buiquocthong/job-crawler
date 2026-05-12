#!/usr/bin/env python3
"""
Vancouver Job Crawler
=====================
Crawl Indeed + Glassdoor cho các vị trí tài chính / data science / C-suite ở Vancouver BC.
Lọc: min $60k/yr hoặc $30/hr | 3 ngày gần nhất | Vancouver area

Cách chạy:
  pip install -r requirements.txt
  python3 vancouver_job_crawler.py        # Chạy thật
  python3 vancouver_job_crawler.py --demo # Demo không cần credentials

Biến môi trường:
  SCRAPER_API_KEY           — ScraperAPI key (bắt buộc để salary coverage cao)
  JOB_PROXY                 — "user:pass@host:port" (thay thế ScraperAPI)
  POWER_AUTOMATE_URL        — MS Teams webhook (tuỳ chọn)
  GDRIVE_CLIENT_ID          — OAuth2 Client ID
  GDRIVE_CLIENT_SECRET      — OAuth2 Client Secret
  GDRIVE_REFRESH_TOKEN      — OAuth2 Refresh Token
  GOOGLE_DRIVE_FOLDER_ID    — ID folder Google Drive

Lấy OAuth2 token lần đầu:
  python3 vancouver_job_crawler.py --get-token
"""

import os, sys, re, json, time, logging
import urllib.parse
from datetime import date, timedelta
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests
from bs4 import BeautifulSoup
from jobspy import scrape_jobs

# ─── CONFIG ──────────────────────────────────────────────────────────────────

SCRAPER_API_KEY    = os.getenv("SCRAPER_API_KEY", "")
POWER_AUTOMATE_URL = os.getenv("POWER_AUTOMATE_URL", "")
GDRIVE_CLIENT_ID   = os.getenv("GDRIVE_CLIENT_ID", "")
GDRIVE_CLIENT_SECRET = os.getenv("GDRIVE_CLIENT_SECRET", "")
GDRIVE_REFRESH_TOKEN = os.getenv("GDRIVE_REFRESH_TOKEN", "")
GDRIVE_FOLDER_ID   = os.getenv("GOOGLE_DRIVE_FOLDER_ID", "")
_PROXY_ENV         = os.getenv("JOB_PROXY", "")

RESULTS_PER_SEARCH = 50
DAYS_OLD           = 3
DISTANCE_KM        = 35
MIN_ANNUAL         = 60_000
MIN_HOURLY         = 30.0
ENRICH_WORKERS     = 5
ENRICH_MAX_JOBS    = 100
OUTPUT_FILE        = Path(f"vancouver_jobs_{date.today()}.csv")

KEYWORD_GROUPS = {
    "finance":    ["Financial Analyst", "FP&A Analyst", "Investment Analyst",
                   "Actuarial Analyst", "CFA Analyst"],
    "data":       ["Data Scientist", "Machine Learning Engineer",
                   "Quantitative Analyst", "Analytics Engineer"],
    "investment": ["CFA Investment", "Actuary", "Investment Management",
                   "Equity Trader", "Portfolio Manager", "Investment Associate"],
    "c_suite":    ["Chief Executive Officer", "Chief Technology Officer",
                   "Chief Information Officer", "Chief AI Officer",
                   "Chief Investment Officer", "CEO", "CTO"],
}

LOCATIONS = [
    "Vancouver, BC, Canada",
    "Burnaby, BC, Canada",
    "North Vancouver, BC, Canada",
]

ALLOWED_CITIES = {
    "vancouver", "burnaby", "north vancouver", "west vancouver",
    "new westminster", "richmond", "coquitlam", "port moody", "remote",
}

OUTPUT_COLS = [
    "title", "company_name", "location_str", "salary_display",
    "min_amount", "max_amount", "interval", "currency", "salary_source",
    "apply_method", "date_posted", "job_url", "job_url_direct",
    "site", "search_keyword", "search_group",
]

# ─── LOGGING ─────────────────────────────────────────────────────────────────

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s",
                    datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

# ─── SALARY PARSING ──────────────────────────────────────────────────────────

def _clean_interval(v) -> str | None:
    lbl = str(v or "").strip().lower().split(".")[-1]  # handle CompensationInterval.YEARLY
    return {
        "yearly": "yearly", "year": "yearly", "yr": "yearly", "annual": "yearly", "annually": "yearly",
        "hourly": "hourly", "hour": "hourly", "hr": "hourly",
        "monthly": "monthly", "month": "monthly",
        "weekly": "weekly", "week": "weekly",
    }.get(lbl)


def _clean_currency(v, location=None) -> str:
    lbl = str(v or "").strip().upper()
    if lbl in ("CAD", "USD"): return lbl
    if location and ("canada" in location.lower() or ", bc" in location.lower()):
        return "CAD"
    return "CAD"


def parse_salary_text(text: str, currency_hint="CAD") -> tuple | None:
    """
    Parse salary từ text string.
    Returns (min, max, interval, currency) hoặc None.

    KEY FIX: Đây là hàm được port từ JS code (parseSalary + salaryQualifies).
    JS code đạt ~100% coverage vì ScraperAPI render=true làm JS của Indeed chạy
    → salary text xuất hiện trong DOM → cheerio.text() lấy được trực tiếp.
    Hàm này xử lý tất cả format phổ biến từ Indeed CA.
    """
    if not text: return None
    text = (text.replace("\u2013", "-").replace("\u2014", "-")
                .replace("\u2212", "-").replace("\xa0", " "))

    _CUR = r"(?:CAD|USD|C\$|US\$|\$)"
    _INT = (r"(?P<interval>per\s*year|a\s*year|/\s*yr?|yr\.?|yearly|annually|annual|"
            r"per\s*hour|an\s*hour|/\s*hr?|hr\.?|hourly|"
            r"per\s*month|a\s*month|/\s*mo|monthly|"
            r"per\s*week|a\s*week|/\s*wk|weekly)")

    patterns = [
        # $80,000 - $100,000 a year  /  $45.00 – $55.00 /hr
        (r"(?P<currency>" + _CUR + r")\s*(?P<min>[\d,]+(?:\.\d+)?)\s*(?P<min_k>[kK]?)"
         r"\s*[-–]\s*(?:" + _CUR + r")?\s*(?P<max>[\d,]+(?:\.\d+)?)\s*(?P<max_k>[kK]?)"
         r"\s*" + _INT + r"?"),
        # $75,000 /year  /  $35/hr
        (r"(?P<currency>" + _CUR + r")\s*(?P<min>[\d,]+(?:\.\d+)?)\s*(?P<min_k>[kK]?)"
         r"\s*" + _INT),
        # From $60,000 a year
        (r"(?:from|up\s+to|starting\s+at)\s+(?P<currency>" + _CUR + r")\s*"
         r"(?P<min>[\d,]+(?:\.\d+)?)\s*(?P<min_k>[kK]?)\s*" + _INT + r"?"),
        # 70,000 to 90,000 per year (no $ sign)
        (r"(?<!\d)(?P<min>\d{2,3},\d{3})\s*(?P<min_k>[kK]?)\s*(?:[-–]|to)\s*"
         r"(?P<max>\d{2,3},\d{3})\s*(?P<max_k>[kK]?)\s*" + _INT + r"?"),
        # Salary: $90,000
        (r"(?:salary|compensation|pay)[:\s]+(?P<currency>" + _CUR + r")?\s*"
         r"(?P<min>[\d,]+(?:\.\d+)?)\s*(?P<min_k>[kK]?)"),
    ]

    def _amount(val, k="") -> float | None:
        if not val: return None
        try:
            n = float(str(val).replace(",", ""))
            return n * 1000 if str(k).lower() == "k" else n
        except ValueError:
            return None

    interval_map = {
        "peryear": "yearly", "ayear": "yearly", "yr": "yearly", "yr.": "yearly",
        "yearly": "yearly", "annually": "yearly", "annual": "yearly",
        "perhour": "hourly", "anhour": "hourly", "hr": "hourly", "hr.": "hourly", "hourly": "hourly",
        "permonth": "monthly", "amonth": "monthly", "mo": "monthly", "monthly": "monthly",
        "perweek": "weekly", "aweek": "weekly", "wk": "weekly", "weekly": "weekly",
    }

    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if not m: continue
        gd = m.groupdict()

        mn = _amount(gd.get("min"), gd.get("min_k", ""))
        mx = _amount(gd.get("max"), gd.get("max_k", ""))
        if not mn or mn <= 0 or mn > 500_000: continue
        if mx and (mx < mn or mx / mn > 10): mx = None

        raw_int = (gd.get("interval") or "").lower().replace(" ", "").replace("/", "")
        interval = interval_map.get(raw_int)
        if not interval:
            interval = "hourly" if mn < 300 else ("monthly" if mn < 1000 else "yearly")

        cur_raw = (gd.get("currency") or "").strip().upper()
        currency = {"C$": "CAD", "US$": "USD", "$": None}.get(cur_raw, cur_raw) or currency_hint

        return int(mn) if mn >= 100 else round(mn, 2), \
               (int(mx) if mx and mx >= 100 else (round(mx, 2) if mx else None)), \
               interval, currency

    return None


def format_salary(row) -> str:
    mn   = row.get("min_amount")
    mx   = row.get("max_amount")
    intv = str(row.get("interval") or "")
    curr = str(row.get("currency") or "CAD")
    src  = str(row.get("salary_source") or "")

    def fmt(v) -> str | None:
        # Phải dùng _isna() thay vì "is None" vì pandas gán None → numpy.float64 NaN
        # float('nan') > 200 = False → f"${float(v):.2f}" → "$nan" (bug!)
        if _isna(v): return None
        f = float(v)
        return f"${int(f):,}" if f > 200 else f"${f:.2f}"

    parts = [p for p in [fmt(mn), fmt(mx)] if p]
    if not parts: return "N/A"

    lbl = {"yearly": "/yr", "hourly": "/hr", "monthly": "/mo", "weekly": "/wk"}.get(intv, "")
    est = " (est.)" if "parse" in src else ""
    return f"{curr} {' - '.join(parts)}{lbl}{est}"


# ─── SALARY RESOLUTION (từ jobspy data) ──────────────────────────────────────

def _isna(v) -> bool:
    if v is None: return True
    try: return bool(pd.isna(v))
    except: return False


def _to_num(v):
    if _isna(v): return None
    try:
        n = float(v)
        return int(n) if n >= 100 else round(n, 2)
    except: return None


def resolve_salary(row) -> tuple:
    """Resolve salary từ jobspy row. Returns (min, max, interval, currency, source)."""
    intv = _clean_interval(row.get("interval"))
    curr = _clean_currency(row.get("currency"), row.get("location"))

    # 1. Direct numeric fields from jobspy
    mn, mx = _to_num(row.get("min_amount")), _to_num(row.get("max_amount"))
    if mn is not None or mx is not None:
        return mn, mx, intv, curr, "direct"

    # 2. Compensation object (jobspy Compensation)
    comp = row.get("compensation")
    if comp is not None and not _isna(comp):
        if isinstance(comp, dict):
            mn = _to_num(comp.get("min_amount") or comp.get("min"))
            mx = _to_num(comp.get("max_amount") or comp.get("max"))
            iv = comp.get("interval") or comp.get("pay_period")
        else:
            mn = _to_num(getattr(comp, "min_amount", None) or getattr(comp, "min", None))
            mx = _to_num(getattr(comp, "max_amount", None) or getattr(comp, "max", None))
            iv = getattr(comp, "interval", None) or getattr(comp, "pay_period", None)
        if hasattr(iv, "value"): iv = iv.value
        if mn is not None or mx is not None:
            return mn, mx, _clean_interval(str(iv)) or intv, curr, "compensation_obj"

    # 3. salary_text / salary fields
    for field in ("salary", "salary_text"):
        parsed = parse_salary_text(str(row.get(field) or ""), curr)
        if parsed: return *parsed, "text_parse"

    # 4. Parse từ description (salary thường xuất hiện trong description)
    desc = str(row.get("description") or "")
    parsed = parse_salary_text(desc, curr)
    if parsed: return *parsed, "desc_parse"

    return None, None, intv, curr, None


# ─── SALARY ENRICHMENT (fetch job page khi salary = N/A) ─────────────────────

def _build_proxies() -> dict | None:
    if not _PROXY_ENV: return None
    url = f"http://{_PROXY_ENV}" if not _PROXY_ENV.startswith("http") else _PROXY_ENV
    return {"http": url, "https": url}


def _fetch_html(url: str) -> str | None:
    """
    Fetch job page HTML.

    KEY FIX: Dùng ScraperAPI với render=true — đây là lý do JS code đạt ~100%
    salary coverage. render=true cho phép JavaScript của Indeed chạy hoàn toàn
    trước khi trả HTML, do đó salary element với data-testid xuất hiện trong DOM.
    Không có render=true, salary thường chỉ nằm trong JS bundle (không parse được
    bằng regex thông thường).
    """
    if SCRAPER_API_KEY:
        try:
            resp = requests.get(
                "https://api.scraperapi.com/",
                params={"api_key": SCRAPER_API_KEY, "url": url,
                        "render": "true", "country_code": "ca"},
                timeout=60,
            )
            if resp.status_code == 200:
                return resp.text
        except Exception as e:
            log.debug(f"ScraperAPI error: {e}")

    # Fallback: direct request với session (có cookie)
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
            "Accept-Language": "en-CA,en;q=0.9",
        }
        resp = requests.get(url, headers=headers, proxies=_build_proxies(), timeout=20)
        if resp.status_code == 200:
            return resp.text
    except Exception as e:
        log.debug(f"Fetch error {url[:60]}: {e}")

    return None


# Các CSS selector salary (port từ JS code — parseSalary dùng cheerio selectors)
_SALARY_SELECTORS = [
    '[data-testid="attribute_snippet_testid"]',
    '[data-testid="salary-snippet"]',
    '[data-testid="salaryInfoAndJobType"]',
    '[data-testid="jobsearch-JobMetadataHeader-salaryInfoAndJobType"]',
    ".salary-snippet-container",
    ".estimated-salary-container",
    '[class*="salaryInfo"]',
    '[class*="salary"]',
    '[class*="Salary"]',
]


def _parse_salary_from_html(html: str, currency_hint="CAD") -> tuple | None:
    """
    Parse salary từ HTML đã fetch. 3 lớp fallback:
    1. BeautifulSoup DOM selectors (hoạt động khi HTML đã JS-rendered)
    2. JSON embedded trong <script> (mosaic-data, JSON-LD)
    3. Text scan toàn trang
    """
    if not html: return None

    # Layer 1: DOM selectors (hiệu quả nhất khi ScraperAPI render=true)
    try:
        soup = BeautifulSoup(html, "lxml")
        for sel in _SALARY_SELECTORS:
            el = soup.select_one(sel)
            if not el: continue
            text = el.get_text(separator=" ", strip=True)
            if "$" in text or "salary" in text.lower():
                parsed = parse_salary_text(text, currency_hint)
                if parsed: return parsed
    except Exception:
        pass

    # Layer 2a: Indeed mosaic-data JSON (raw HTML)
    m = re.search(r'<script\s+id=["\']mosaic-data["\'][^>]*>(.*?)</script>',
                  html, re.DOTALL | re.IGNORECASE)
    if m:
        try:
            mosaic = json.loads(m.group(1))
            json_str = json.dumps(mosaic)
            # extractedSalary numeric
            es_m = re.search(
                r'"extractedSalary"\s*:\s*\{"min"\s*:\s*([\d.]+)[^}]*"max"\s*:\s*([\d.]+)[^}]*"type"\s*:\s*"([^"]+)"',
                json_str, re.IGNORECASE
            )
            if es_m:
                mn, mx = float(es_m.group(1)), float(es_m.group(2))
                intv = "hourly" if "hour" in es_m.group(3).lower() else "yearly"
                return int(min(mn, mx)), int(max(mn, mx)), intv, currency_hint
            # salarySnippet / formattedSalary text
            for key in ("salarySnippet", "formattedSalary", "salaryText", "salaryRange"):
                sm = re.search(rf'"{key}"\s*:\s*"([^"{{}}]{{5,100}})"', json_str, re.IGNORECASE)
                if sm:
                    text = sm.group(1).replace("\\u2013", "-")
                    parsed = parse_salary_text(text, currency_hint)
                    if parsed: return parsed
        except Exception:
            pass

    # Layer 2b: JSON-LD schema.org
    for m in re.finditer(r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
                          html, re.DOTALL | re.IGNORECASE):
        try:
            data = json.loads(m.group(1))
            items = data if isinstance(data, list) else [data]
            for item in items:
                bs = item.get("baseSalary") or item.get("estimatedSalary")
                if not bs or not isinstance(bs, dict): continue
                val = bs.get("value") or {}
                if isinstance(val, dict):
                    text = f"${val.get('minValue','')} - ${val.get('maxValue','')} {val.get('unitText','')}".strip()
                elif isinstance(val, (int, float, str)):
                    text = f"${val} {bs.get('unitText','')}".strip()
                else: continue
                parsed = parse_salary_text(text, currency_hint)
                if parsed: return parsed
        except Exception:
            pass

    # Layer 3: Full text scan
    text_only = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", html[:30_000]))
    return parse_salary_text(text_only, currency_hint)


def enrich_salaries(df: pd.DataFrame) -> pd.DataFrame:
    """Fetch job pages để lấy salary cho các job N/A."""
    mask  = df["salary_display"] == "N/A"
    na_df = df[mask].head(ENRICH_MAX_JOBS)
    if na_df.empty:
        log.info("Không có job N/A cần enrich."); return df

    log.info(f"🔍 Enriching {len(na_df)} jobs "
             f"({'ScraperAPI=ON' if SCRAPER_API_KEY else 'no ScraperAPI — coverage thấp'})...")

    def _enrich_one(idx_row):
        idx, row = idx_row
        curr = str(row.get("currency") or "CAD")
        for url in [row.get("job_url_direct"), row.get("job_url")]:
            url_str = str(url or "").strip()
            if not url_str.startswith("http"): continue
            time.sleep(0.5)
            html = _fetch_html(url_str)
            result = _parse_salary_from_html(html, curr)
            if result: return idx, result
        return idx, None

    enriched = 0
    with ThreadPoolExecutor(max_workers=ENRICH_WORKERS) as ex:
        futures = {ex.submit(_enrich_one, (idx, row)): idx for idx, row in na_df.iterrows()}
        for i, fut in enumerate(as_completed(futures), 1):
            idx, result = fut.result()
            if result:
                mn, mx, intv, curr = result
                df.at[idx, "min_amount"]    = mn
                df.at[idx, "max_amount"]    = mx
                df.at[idx, "interval"]      = intv
                df.at[idx, "currency"]      = curr
                df.at[idx, "salary_source"] = "enriched"
                df.at[idx, "salary_display"] = format_salary(df.loc[idx])
                enriched += 1
                log.info(f"  ✅ [{i}/{len(na_df)}] {df.at[idx,'title'][:30]} → {df.at[idx,'salary_display']}")

    log.info(f"✅ Enrich xong: +{enriched}/{len(na_df)} jobs có lương")
    if enriched == 0:
        log.warning("⚠️  Enrich = 0 — Indeed block direct requests. Cần SCRAPER_API_KEY hoặc JOB_PROXY.")
    return df


# ─── CRAWL ────────────────────────────────────────────────────────────────────

def crawl_jobs() -> pd.DataFrame:
    all_frames = []
    total = sum(len(v) for v in KEYWORD_GROUPS.values()) * len(LOCATIONS)
    done  = 0

    for group, keywords in KEYWORD_GROUPS.items():
        for keyword in keywords:
            for location in LOCATIONS:
                done += 1
                log.info(f"[{done}/{total}] '{keyword}' @ {location}")
                try:
                    df = scrape_jobs(
                        site_name=["indeed"],
                        search_term=keyword, location=location,
                        country_indeed="canada",
                        results_wanted=RESULTS_PER_SEARCH,
                        hours_old=DAYS_OLD * 24, distance=DISTANCE_KM,
                        description_format="markdown",
                        proxies=[_PROXY_ENV] if _PROXY_ENV else None,
                        verbose=0,
                    )
                    if df is not None and not df.empty:
                        df["search_keyword"] = keyword
                        df["search_group"]   = group
                        log.info(f"  → {len(df)} jobs | salary_direct={int(df['min_amount'].notna().sum()) if 'min_amount' in df.columns else 0}")
                        all_frames.append(df)
                    else:
                        log.warning("  → 0 jobs")
                except Exception as e:
                    log.warning(f"  → Lỗi: {e}")
                time.sleep(2)

    if not all_frames:
        log.error("Không lấy được job nào. Kiểm tra JOB_PROXY / SCRAPER_API_KEY.")
        return pd.DataFrame()

    raw = pd.concat(all_frames, ignore_index=True)
    log.info(f"Tổng thô (có trùng): {len(raw)}")
    return raw


# ─── NORMALIZE ────────────────────────────────────────────────────────────────

def normalize(df: pd.DataFrame) -> pd.DataFrame:
    # Đảm bảo các cột cần thiết tồn tại
    for col in ["min_amount", "max_amount", "interval", "currency", "date_posted",
                "listing_type", "job_url_direct", "location", "compensation",
                "salary", "salary_text", "description", "salary_source"]:
        if col not in df.columns: df[col] = None

    # location object → string
    def loc_str(v):
        if isinstance(v, str): return v
        try: return ", ".join(filter(None, [getattr(v, "city", None), getattr(v, "state", None)]))
        except: return str(v or "")
    df["location_str"] = df["location"].apply(loc_str)

    # Resolve salary fields
    sal = df.apply(resolve_salary, axis=1, result_type="expand")
    sal.columns = ["min_amount", "max_amount", "interval", "currency", "salary_source"]
    df[["min_amount", "max_amount", "interval", "currency", "salary_source"]] = sal

    df["salary_display"] = df.apply(format_salary, axis=1)
    df["date_posted"]    = pd.to_datetime(df["date_posted"], errors="coerce").dt.date
    df["apply_method"]   = df.apply(
        lambda r: "Apply on Company Site" if str(r.get("job_url_direct") or "").startswith("http")
        else ("Apply Now" if "apply_direct" in str(r.get("listing_type") or "").lower()
              else "Apply on Company Site"), axis=1
    )

    has_sal = int(df["min_amount"].notna().sum())
    log.info(f"Salary coverage: {has_sal}/{len(df)} ({has_sal/len(df)*100:.0f}%)")
    return df


# ─── FILTER ───────────────────────────────────────────────────────────────────

def filter_jobs(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty: return df
    n0 = len(df)

    df = df.drop_duplicates(subset=["job_url"], keep="first")
    log.info(f"Sau dedup URL    : {len(df)} (bỏ {n0 - len(df)})")

    df = df[df["location_str"].apply(
        lambda l: any(c in l.lower() for c in ALLOWED_CITIES) if l else False)]
    log.info(f"Sau lọc địa điểm : {len(df)}")

    cutoff = date.today() - timedelta(days=DAYS_OLD)
    df = df[df["date_posted"].apply(
        lambda d: d is None or (isinstance(d, date) and d >= cutoff))]
    log.info(f"Sau lọc ngày     : {len(df)}")

    min_by_interval = {"yearly": MIN_ANNUAL, "hourly": MIN_HOURLY,
                       "monthly": MIN_ANNUAL / 12, "weekly": MIN_HOURLY * 40}

    def salary_ok(row) -> bool:
        mn, mx = row.get("min_amount"), row.get("max_amount")
        if _isna(mn) and _isna(mx): return True  # giữ lại để enrich
        v = float(mn if not _isna(mn) else mx)
        return v >= min_by_interval.get(str(row.get("interval") or ""), MIN_ANNUAL)

    df = df[df.apply(salary_ok, axis=1)]
    log.info(f"Sau lọc lương    : {len(df)}")
    return df.reset_index(drop=True)


# ─── SAVE ─────────────────────────────────────────────────────────────────────

def save_csv(df: pd.DataFrame) -> Path:
    cols = [c for c in OUTPUT_COLS if c in df.columns]
    out  = df[cols].sort_values(["min_amount", "date_posted"],
                                ascending=[False, False], na_position="last")
    out.to_csv(OUTPUT_FILE, index=False, encoding="utf-8-sig")
    log.info(f"Saved {len(out)} jobs → {OUTPUT_FILE.resolve()}")
    return OUTPUT_FILE


def save_excel(df: pd.DataFrame) -> Path | None:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl chưa cài — bỏ qua. pip install openpyxl"); return None

    xlsx_path = OUTPUT_FILE.with_suffix(".xlsx")
    cols = [c for c in OUTPUT_COLS if c in df.columns]
    out  = df[cols].sort_values(["min_amount", "date_posted"],
                                ascending=[False, False], na_position="last")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vancouver Jobs"

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(border_style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill, cell.font = hdr_fill, hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = bdr

    na_fill  = PatternFill("solid", fgColor="FFE0E0")
    sal_col  = (cols.index("salary_display") + 1) if "salary_display" in cols else None
    col_widths = {"title": 36, "company_name": 24, "location_str": 22,
                  "salary_display": 30, "apply_method": 24, "date_posted": 13,
                  "job_url": 50, "job_url_direct": 50, "search_keyword": 28}

    for ri, row_data in enumerate(out.itertuples(index=False), 2):
        is_na = sal_col and str(getattr(row_data, "salary_display", "")) == "N/A"
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = bdr
            cell.alignment = Alignment(vertical="center")
            if is_na and ci == sal_col:
                cell.fill = na_fill

    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 14)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(xlsx_path)
    log.info(f"Saved Excel → {xlsx_path.resolve()}")
    return xlsx_path


# ─── GOOGLE DRIVE UPLOAD ──────────────────────────────────────────────────────

def upload_to_drive(file_path: Path) -> tuple[str | None, str | None]:
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request as GoogleRequest
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaFileUpload
    except ImportError:
        log.warning("google-api-python-client chưa cài — bỏ qua Drive upload."); return None, None

    if not all([GDRIVE_CLIENT_ID, GDRIVE_CLIENT_SECRET, GDRIVE_REFRESH_TOKEN, GDRIVE_FOLDER_ID]):
        log.warning("⚠️  Thiếu GDRIVE credentials — bỏ qua Drive upload."); return None, None

    try:
        creds = Credentials(
            token=None, refresh_token=GDRIVE_REFRESH_TOKEN,
            token_uri="https://oauth2.googleapis.com/token",
            client_id=GDRIVE_CLIENT_ID, client_secret=GDRIVE_CLIENT_SECRET,
            scopes=["https://www.googleapis.com/auth/drive"],
        )
        if not creds.valid: creds.refresh(GoogleRequest())
        service = build("drive", "v3", credentials=creds)

        # Xoá file cũ cùng tên
        existing = service.files().list(
            q=f"name='{file_path.name}' and '{GDRIVE_FOLDER_ID}' in parents and trashed=false",
            fields="files(id)"
        ).execute()
        for f in existing.get("files", []):
            service.files().delete(fileId=f["id"]).execute()

        suffix = file_path.suffix.lower()
        mime = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                if suffix == ".xlsx" else "text/csv")
        uploaded = service.files().create(
            body={"name": file_path.name, "parents": [GDRIVE_FOLDER_ID]},
            media_body=MediaFileUpload(str(file_path), mimetype=mime, resumable=True),
            fields="id,name"
        ).execute()
        file_id = uploaded["id"]
        service.permissions().create(fileId=file_id, body={"type": "anyone", "role": "reader"}).execute()
        log.info(f"✅ Drive upload: {uploaded['name']} (id={file_id})")
        return f"https://drive.google.com/drive/folders/{GDRIVE_FOLDER_ID}", \
               f"https://drive.google.com/file/d/{file_id}/view"

    except Exception as e:
        log.error(f"❌ Drive upload lỗi: {e}"); return None, None


# ─── NOTIFY ───────────────────────────────────────────────────────────────────

def send_teams_notification(df: pd.DataFrame, folder_url: str | None, file_url: str | None):
    if not POWER_AUTOMATE_URL:
        log.warning("POWER_AUTOMATE_URL chưa cấu hình — bỏ qua Teams."); return

    has_sal = int(df["salary_display"].ne("N/A").sum()) if "salary_display" in df.columns else 0
    download_url = file_url or folder_url or ""
    card = {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard", "version": "1.3",
        "body": [
            {"type": "TextBlock", "text": f"Vancouver Jobs — {date.today():%d/%m/%Y}",
             "weight": "Bolder", "size": "Medium"},
            {"type": "FactSet", "facts": [
                {"title": "Tổng jobs:",  "value": str(len(df))},
                {"title": "Có lương:",   "value": f"{has_sal}/{len(df)}"},
                {"title": "Nguồn:",      "value": "Indeed + Glassdoor CA"},
            ]},
        ],
        "actions": ([
            {"type": "Action.OpenUrl", "title": "Mở Excel", "url": download_url},
            {"type": "Action.OpenUrl", "title": "Mở Drive Folder", "url": folder_url or ""},
        ] if download_url else []),
    }
    try:
        resp = requests.post(POWER_AUTOMATE_URL, json={"card": card}, timeout=30)
        log.info("✅ Đã gửi Teams." if resp.status_code in (200, 202)
                 else f"❌ Power Automate {resp.status_code}: {resp.text[:200]}")
    except Exception as e:
        log.error(f"❌ Teams error: {e}")


# ─── DEMO ─────────────────────────────────────────────────────────────────────

def make_demo_data() -> pd.DataFrame:
    today = date.today()
    rows = [
        ("Financial Analyst",       "Telus",           "Vancouver, BC",  70000, 90000,  "yearly"),
        ("Data Scientist",          "Hootsuite",        "Vancouver, BC",  90000, 115000, "yearly"),
        ("ML Engineer",             "Microsoft Canada", "Vancouver, BC", 105000, 145000, "yearly"),
        ("Portfolio Manager",       "Mackenzie Invest.","Vancouver, BC", 130000, 175000, "yearly"),
        ("Chief Technology Officer","Visier",           "Vancouver, BC", 200000, 280000, "yearly"),
        ("Investment Analyst",      "RBC Wealth Mgmt",  "Vancouver, BC",  80000, 105000, "yearly"),
        ("Equity Trader",           "Canaccord Genuity","Vancouver, BC", 100000, 160000, "yearly"),
        ("Data Analyst (hourly)",   "BCAA",           "North Vancouver",   35.0,   45.0, "hourly"),
        # Bị lọc — lương thấp
        ("Junior Analyst",         "Small Co",        "Vancouver, BC",  40000,  55000, "yearly"),
        # Bị lọc — sai địa điểm
        ("Financial Analyst",      "TD Bank Toronto", "Toronto, ON",    80000, 100000, "yearly"),
    ]
    records = []
    for i, (title, co, loc, mn, mx, intv) in enumerate(rows):
        records.append({
            "title": title, "company_name": co, "location": loc,
            "min_amount": mn, "max_amount": mx, "interval": intv,
            "job_url": f"https://ca.indeed.com/viewjob?jk=demo{i:04d}",
            "job_url_direct": "", "date_posted": today - timedelta(days=i % 3),
            "site": "indeed" if i % 2 == 0 else "glassdoor",
            "is_remote": False, "search_keyword": title.split()[0],
            "search_group": "demo", "currency": "CAD", "listing_type": "organic",
        })
    log.info(f"[DEMO] Tạo {len(records)} jobs mẫu")
    return pd.DataFrame(records)


# ─── GET-TOKEN ────────────────────────────────────────────────────────────────

def get_oauth2_token():
    """Lấy OAuth2 refresh token lần đầu (chạy local, không cần trong CI)."""
    if not all([GDRIVE_CLIENT_ID, GDRIVE_CLIENT_SECRET]):
        print("❌ Cần set GDRIVE_CLIENT_ID và GDRIVE_CLIENT_SECRET trước."); return

    auth_url = (
        "https://accounts.google.com/o/oauth2/auth?"
        + urllib.parse.urlencode({
            "client_id": GDRIVE_CLIENT_ID, "redirect_uri": "urn:ietf:wg:oauth:2.0:oob",
            "response_type": "code", "scope": "https://www.googleapis.com/auth/drive",
            "access_type": "offline", "prompt": "consent",
        })
    )
    print(f"\nMở URL này trong trình duyệt:\n{auth_url}\n")
    code = input("Paste authorization code: ").strip()

    resp = requests.post("https://oauth2.googleapis.com/token", data={
        "client_id": GDRIVE_CLIENT_ID, "client_secret": GDRIVE_CLIENT_SECRET,
        "code": code, "grant_type": "authorization_code",
        "redirect_uri": "urn:ietf:wg:oauth:2.0:oob",
    })
    data = resp.json()
    refresh_token = data.get("refresh_token")
    if not refresh_token:
        print(f"❌ Lỗi: {data}"); return

    print(f"\n✅ REFRESH TOKEN:\n{refresh_token}")
    print("\nCopy token trên vào GitHub Secret: GDRIVE_REFRESH_TOKEN")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    demo_mode = "--demo" in sys.argv
    if "--get-token" in sys.argv: get_oauth2_token(); return

    print("\n" + "═" * 60)
    print(f"  Vancouver Job Crawler {'[DEMO MODE]' if demo_mode else ''}")
    print(f"  Ngày    : {DAYS_OLD} ngày gần nhất | Lương: ≥${MIN_ANNUAL:,}/yr hoặc ${MIN_HOURLY}/hr")
    print(f"  Scraper : {'✅ ScraperAPI' if SCRAPER_API_KEY else '⚠️  chưa cấu hình (salary coverage thấp)'}")
    print(f"  Proxy   : {'✅ ' + _PROXY_ENV[:30] if _PROXY_ENV else '⚠️  chưa cấu hình'}")
    print("═" * 60 + "\n")

    raw_df = make_demo_data() if demo_mode else crawl_jobs()
    if raw_df.empty: return

    df = normalize(raw_df)
    df = filter_jobs(df)

    if not demo_mode:
        df = enrich_salaries(df)
        df["salary_display"] = df.apply(format_salary, axis=1)

    csv_path  = save_csv(df)
    xlsx_path = save_excel(df)

    has_sal = int(df["salary_display"].ne("N/A").sum())
    print(f"\n{'═'*60}")
    print(f"  Tổng jobs : {len(df)}")
    print(f"  Có lương  : {has_sal}/{len(df)} ({has_sal/len(df)*100:.0f}%)" if len(df) > 0 else "  Không có kết quả")
    print(f"  File CSV  : {csv_path.resolve()}")
    print("═" * 60 + "\n")

    folder_url, file_url = None, None
    if xlsx_path:
        folder_url, file_url = upload_to_drive(xlsx_path)
    if not file_url:
        folder_url, file_url = upload_to_drive(csv_path)

    send_teams_notification(df, folder_url, file_url)


if __name__ == "__main__":
    main()